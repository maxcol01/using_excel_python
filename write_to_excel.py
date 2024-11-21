import openpyxl # type: ignore
from pathlib import Path

cd = Path.cwd()
WORKBOOK = "WriteToExcel.xlsx"
path_to_file = cd / WORKBOOK

if path_to_file.exists():
    wb = openpyxl.load_workbook(WORKBOOK)
    print("Workbook loaded !")
else:
    wb = openpyxl.Workbook(path_to_file)
    wb.save(WORKBOOK)

data = [
	["Product", "Category", "Price"],
	["Apple Iphone 12", "Electronics", 799],
	["Nike Air Force 1", "Footwear", 90],
	["The Alchemist", "Book", 16.99],
	["Instant Pot Duo", "Home & Kitchen", 89]

]

sheets = wb.sheetnames
ws = wb[sheets[0]]

num_rows = len(data)

rgn_adress = f"A1:C{num_rows}" # there are three column and we want the number of rows to be dynamic !

rng = ws[rgn_adress] # we have now an iterable and we can iterate to populate the cells with data !

for i in range(0, len(data)):
    row = data[i]
    for j in range(0, len(row)):
        val = row[j]
        rng[i][j].value = val

wb.save(path_to_file)