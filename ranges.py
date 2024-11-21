import openpyxl #type: ignore
from pathlib import Path

cd = Path.cwd()
WORBOOK = "Employee Ratings.xlsx"

path_to_file = cd / WORBOOK

if not path_to_file.exists():
    print("Please provide us with the data !")

else:
    wb = openpyxl.load_workbook(path_to_file)


ws = wb[wb.sheetnames[0]]
rng = ws["B2:E11"]

for item in rng:
    for cell in item:
        if isinstance(cell.value, str):
            cell.value = ""
        else:
            if cell.value < 0:
                cell.value = 0
            elif cell.value > 10:
                cell.value = 10
            else:
                continue

wb.save(path_to_file)    


