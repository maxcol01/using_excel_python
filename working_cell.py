import openpyxl #type: ignore
from pathlib import Path


def save_wb(wb, file_path):
    wb.save(file_path)
    print("Workbook saved !")

# Create a workbook for this example ! 
cd = Path.cwd()
WORKBOOK = "WorkingWithCells.xlsx"
file_path = cd / WORKBOOK

if not file_path.exists():
    wb = openpyxl.Workbook()
    wb.save(file_path)
else:
    wb = openpyxl.load_workbook(WORKBOOK)
    print("File already existing, please Carry on !")


ws = wb[wb.sheetnames[0]]
ws["A1"] = "Hello World"
string = ws["A1"].value
reversed_string = string[::-1]

ws["A2"] = reversed_string
ws.cell(row=11, column=11, value="Geetings human")
print(ws.cell(row=11, column=11).value) # since we watch the value we loose the value argument in the .cell() method but we need the .value attribute to display the value
save_wb(wb, WORKBOOK)
