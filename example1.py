
import openpyxl
from pathlib import Path

import openpyxl.worksheet

cd: Path = Path(".")
SPREAD_SHEET_NAME: str = "my_excel_spreadsheet.xlsx"

status: str = "read "

if status == "create":
    wb:  openpyxl.Workbook = openpyxl.Workbook()
    wb.save(cd / f"{SPREAD_SHEET_NAME}")
else:
    wb = openpyxl.load_workbook(cd / f"{SPREAD_SHEET_NAME}")
    
ws: openpyxl.worksheet = wb.active
print(type(ws))

ws["A1"] = "Hello Excel"

wb.save(cd / f"{SPREAD_SHEET_NAME}")