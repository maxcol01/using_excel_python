import openpyxl # type: ignore

WORKBOOK: str = "Employees.xlsx"

wb: openpyxl.Workbook = openpyxl.load_workbook(WORKBOOK)

names: list = ["John Smith", "Amit Patel", "Robert Brown", "Sanaa Al Farsi", "Michael Miller"]

for name in names:
    wb.create_sheet(name)

wb.save(WORKBOOK)

sheet_names: list = wb.sheetnames

to_delete: list = list()
for sheet_name in sheet_names:
    if "Sheet" in sheet_name:
        to_delete.append(sheet_name)

for sheet in range(1, len(to_delete)+1):
    del wb[f"Sheet{sheet}"]

wb.save(WORKBOOK)